from __future__ import annotations

import json
import logging
import re
import sqlite3
from contextlib import contextmanager
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from src.config import Settings, load_settings
from src.gemini_api import GeminiClient, extract_gemini_text
from src.openai_api import OpenAIResponsesClient, extract_json_document, extract_output_text
from src.search.queries import generate_name_variants, normalize_name

log = logging.getLogger("istari.mapping_low_confidence")

_MARKDOWN_LINK_RE = re.compile(r"\[([^\]]+)\]\((https?://[^)\s]+)\)")
_PLAIN_URL_RE = re.compile(r"(?<!\()(?P<url>https?://[^\s)>\]]+)")
_SENTENCE_SPLIT_RE = re.compile(r"(?<=[.!?])\s+")
_LOW_CONFIDENCE_OVERLAY_SOURCE_NAMES = (
    "mapping_links.sqlite",
    "mapping_links.signatory-clean.sqlite",
    "mapping_links.guardian-normalize.sqlite",
    "mapping_links.normalize-after-rest.sqlite",
)
_OTHER_ORGANISATION_HINTS = (
    "open letter",
    "letter",
    "statement",
    "campaign",
    "petition",
    "manifesto",
    "declaration",
    "appeal",
)
_ORGANISATION_SUFFIX_TOKENS = {
    "ltd",
    "limited",
    "plc",
    "llp",
    "inc",
    "incorporated",
    "corp",
    "corporation",
    "company",
    "co",
}
_ORGANISATION_AI_MIN_RATIO = 0.84
_ORGANISATION_AI_MIN_OVERLAP = 2
_ORGANISATION_AI_MAX_CANDIDATES = 5


def _ensure_text_column(
    connection: sqlite3.Connection,
    *,
    table_name: str,
    column_name: str,
    default_value: str = "",
) -> None:
    columns = {
        str(row["name"])
        for row in connection.execute(f"PRAGMA table_info({table_name})").fetchall()
    }
    if column_name in columns:
        return
    escaped = str(default_value).replace("'", "''")
    connection.execute(
        f"ALTER TABLE {table_name} ADD COLUMN {column_name} TEXT NOT NULL DEFAULT '{escaped}'"
    )


def normalize_mapping_label(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _clean_mapping_scalar(value).lower()).strip()


def slugify_mapping_label(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", normalize_mapping_label(value))
    return slug.strip("_") or "node"


def default_mapping_db_path(project_root: Path) -> Path:
    return project_root / "data" / "mapping_links.sqlite"


def default_overlay_mapping_db_path(project_root: Path) -> Path:
    return project_root / "data" / "mapping_links.combined.sqlite"


def _mapping_db_has_rows(database_path: Path) -> bool:
    path = Path(database_path)
    if not path.exists():
        return False
    connection = sqlite3.connect(path)
    try:
        tables = {
            str(row[0])
            for row in connection.execute("SELECT name FROM sqlite_master WHERE type = 'table'").fetchall()
        }
        if "mapping_links" not in tables:
            return False
        return bool(connection.execute("SELECT 1 FROM mapping_links LIMIT 1").fetchone())
    finally:
        connection.close()


def overlay_mapping_source_paths(project_root: Path) -> list[Path]:
    data_dir = Path(project_root) / "data"
    seen: set[Path] = set()
    paths: list[Path] = []
    for name in _LOW_CONFIDENCE_OVERLAY_SOURCE_NAMES:
        path = data_dir / name
        if path in seen or not _mapping_db_has_rows(path):
            continue
        seen.add(path)
        paths.append(path)
    return paths


def rebuild_overlay_mapping_db(project_root: Path) -> Path:
    target_path = default_overlay_mapping_db_path(project_root)
    target_store = MappingStore(target_path)
    target_store.init_db()
    target_store.clear_all()

    source_paths = overlay_mapping_source_paths(project_root)
    if not source_paths:
        return target_path

    seen_entity_keys: set[tuple[str, str, int]] = set()
    seen_link_keys: dict[tuple[str, str, int], int] = {}
    seen_evidence_keys: set[tuple[str, str, int, int]] = set()

    with target_store.managed_connection() as target_connection:
        for source_path in source_paths:
            source_connection = sqlite3.connect(source_path)
            source_connection.row_factory = sqlite3.Row
            try:
                source_dir_row = source_connection.execute(
                    "SELECT source_dir FROM mapping_imports ORDER BY id LIMIT 1"
                ).fetchone()
                source_dir = (
                    str(source_dir_row["source_dir"]).strip()
                    if source_dir_row and str(source_dir_row["source_dir"]).strip()
                    else str(source_path)
                )
                import_id = int(
                    target_connection.execute(
                        "INSERT INTO mapping_imports(source_dir) VALUES(?)",
                        (source_dir,),
                    ).lastrowid
                )
                for row in source_connection.execute(
                    """
                    SELECT *
                    FROM mapping_entities
                    ORDER BY workbook_name, sheet_name, row_number
                    """
                ).fetchall():
                    entity_key = (
                        str(row["workbook_name"]),
                        str(row["sheet_name"]),
                        int(row["row_number"]),
                    )
                    if entity_key in seen_entity_keys:
                        continue
                    seen_entity_keys.add(entity_key)
                    target_connection.execute(
                        """
                        INSERT INTO mapping_entities(
                            import_id,
                            workbook_name,
                            sheet_name,
                            row_number,
                            label,
                            normalized_label,
                            entity_type,
                            description,
                            raw_row_json
                        ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            import_id,
                            row["workbook_name"],
                            row["sheet_name"],
                            row["row_number"],
                            row["label"],
                            row["normalized_label"],
                            row["entity_type"],
                            row["description"],
                            row["raw_row_json"],
                        ),
                    )
                for row in source_connection.execute(
                    """
                    SELECT *
                    FROM mapping_links
                    ORDER BY workbook_name, sheet_name, row_number
                    """
                ).fetchall():
                    link_key = (
                        str(row["workbook_name"]),
                        str(row["sheet_name"]),
                        int(row["row_number"]),
                    )
                    if link_key in seen_link_keys:
                        continue
                    link_id = int(
                        target_connection.execute(
                            """
                            INSERT INTO mapping_links(
                                import_id,
                                workbook_name,
                                sheet_name,
                                row_number,
                                from_label,
                                from_normalized_label,
                                to_label,
                                to_normalized_label,
                                link_type,
                                description,
                                quality,
                                raw_row_json
                            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                import_id,
                                row["workbook_name"],
                                row["sheet_name"],
                                row["row_number"],
                                row["from_label"],
                                row["from_normalized_label"],
                                row["to_label"],
                                row["to_normalized_label"],
                                row["link_type"],
                                row["description"],
                                row["quality"],
                                row["raw_row_json"],
                            ),
                        ).lastrowid
                    )
                    seen_link_keys[link_key] = link_id
                for row in source_connection.execute(
                    """
                    SELECT
                        link.workbook_name,
                        link.sheet_name,
                        link.row_number,
                        evidence.ordinal,
                        evidence.evidence_kind,
                        evidence.title,
                        evidence.url,
                        evidence.snippet,
                        evidence.document_summary
                    FROM mapping_evidence AS evidence
                    JOIN mapping_links AS link ON link.id = evidence.mapping_link_id
                    ORDER BY link.workbook_name, link.sheet_name, link.row_number, evidence.ordinal
                    """
                ).fetchall():
                    evidence_key = (
                        str(row["workbook_name"]),
                        str(row["sheet_name"]),
                        int(row["row_number"]),
                        int(row["ordinal"]),
                    )
                    if evidence_key in seen_evidence_keys:
                        continue
                    new_link_id = seen_link_keys.get(evidence_key[:3])
                    if not new_link_id:
                        continue
                    seen_evidence_keys.add(evidence_key)
                    target_connection.execute(
                        """
                        INSERT INTO mapping_evidence(
                            mapping_link_id,
                            ordinal,
                            evidence_kind,
                            title,
                            url,
                            snippet,
                            document_summary
                        ) VALUES(?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            new_link_id,
                            row["ordinal"],
                            row["evidence_kind"],
                            row["title"],
                            row["url"],
                            row["snippet"],
                            row["document_summary"],
                        ),
                    )
            finally:
                source_connection.close()

    return target_path


def _clean_mapping_scalar(value: str) -> str:
    cleaned = str(value or "").replace("\ufffd", "")
    return re.sub(r"\s+", " ", cleaned).strip()


def _clean_mapping_block(value: str) -> str:
    cleaned = str(value or "").replace("\ufffd", "")
    return cleaned.strip()


def canonicalize_entity_type(value: str) -> str:
    text = _clean_mapping_scalar(value).lower()
    if text in {"person", "individual"}:
        return "individual"
    if text in {"organisation", "organization", "org"}:
        return "organisation"
    return text


def canonicalize_link_type(value: str) -> str:
    return _clean_mapping_scalar(value).lower()


def summarize_mapping_text(value: str, *, max_chars: int = 220) -> str:
    raw_text = _clean_mapping_block(value)
    if not raw_text:
        return ""
    text = _MARKDOWN_LINK_RE.sub(lambda match: match.group(1).strip(), raw_text)
    text = _PLAIN_URL_RE.sub("", text)
    text = re.sub(r"\s+", " ", text).strip(" -:;,.")
    if not text:
        return ""
    first_sentence = _SENTENCE_SPLIT_RE.split(text, maxsplit=1)[0].strip()
    if first_sentence and len(first_sentence) <= max_chars:
        return first_sentence
    if len(text) <= max_chars:
        return text
    shortened = text[: max_chars - 3].rsplit(" ", 1)[0].strip()
    return f"{shortened or text[: max_chars - 3].strip()}..."


def extract_evidence_links(text: str) -> list[dict[str, str]]:
    raw_text = str(text or "")
    found: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()

    for title, url in _MARKDOWN_LINK_RE.findall(raw_text):
        key = (url.strip(), title.strip())
        if key in seen:
            continue
        seen.add(key)
        found.append(
            {
                "kind": "markdown_link",
                "title": title.strip() or url.strip(),
                "url": url.strip(),
            }
        )

    masked_text = _MARKDOWN_LINK_RE.sub("", raw_text)
    for match in _PLAIN_URL_RE.finditer(masked_text):
        url = match.group("url").rstrip(".,);:")
        key = (url, "")
        if not url or key in seen:
            continue
        seen.add(key)
        found.append(
            {
                "kind": "plain_url",
                "title": url,
                "url": url,
            }
        )

    return found


class MappingStore:
    def __init__(self, database_path: Path) -> None:
        self.database_path = Path(database_path)
        self.database_path.parent.mkdir(parents=True, exist_ok=True)

    def connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.database_path)
        connection.row_factory = sqlite3.Row
        return connection

    @contextmanager
    def managed_connection(self):
        connection = self.connect()
        try:
            yield connection
            connection.commit()
        finally:
            connection.close()

    def init_db(self) -> None:
        with self.managed_connection() as connection:
            connection.executescript(
                """
                PRAGMA foreign_keys = ON;

                CREATE TABLE IF NOT EXISTS mapping_imports (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source_dir TEXT NOT NULL,
                    imported_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );

                CREATE TABLE IF NOT EXISTS mapping_entities (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    import_id INTEGER NOT NULL REFERENCES mapping_imports(id) ON DELETE CASCADE,
                    workbook_name TEXT NOT NULL,
                    sheet_name TEXT NOT NULL,
                    row_number INTEGER NOT NULL,
                    label TEXT NOT NULL,
                    normalized_label TEXT NOT NULL,
                    entity_type TEXT NOT NULL DEFAULT '',
                    description TEXT NOT NULL DEFAULT '',
                    raw_row_json TEXT NOT NULL,
                    UNIQUE(workbook_name, sheet_name, row_number)
                );

                CREATE TABLE IF NOT EXISTS mapping_links (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    import_id INTEGER NOT NULL REFERENCES mapping_imports(id) ON DELETE CASCADE,
                    workbook_name TEXT NOT NULL,
                    sheet_name TEXT NOT NULL,
                    row_number INTEGER NOT NULL,
                    from_label TEXT NOT NULL,
                    from_normalized_label TEXT NOT NULL,
                    to_label TEXT NOT NULL,
                    to_normalized_label TEXT NOT NULL,
                    link_type TEXT NOT NULL DEFAULT '',
                    description TEXT NOT NULL DEFAULT '',
                    quality TEXT NOT NULL DEFAULT 'low',
                    raw_row_json TEXT NOT NULL,
                    UNIQUE(workbook_name, sheet_name, row_number)
                );

                CREATE TABLE IF NOT EXISTS mapping_evidence (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    mapping_link_id INTEGER NOT NULL REFERENCES mapping_links(id) ON DELETE CASCADE,
                    ordinal INTEGER NOT NULL,
                    evidence_kind TEXT NOT NULL DEFAULT 'plain_url',
                    title TEXT NOT NULL DEFAULT '',
                    url TEXT NOT NULL DEFAULT '',
                    snippet TEXT NOT NULL DEFAULT '',
                    UNIQUE(mapping_link_id, ordinal)
                );

                CREATE TABLE IF NOT EXISTS mapping_matches (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    mapping_link_id INTEGER NOT NULL REFERENCES mapping_links(id) ON DELETE CASCADE,
                    endpoint TEXT NOT NULL,
                    run_key TEXT NOT NULL,
                    matched_node_id TEXT NOT NULL,
                    matched_node_label TEXT NOT NULL,
                    match_type TEXT NOT NULL,
                    UNIQUE(mapping_link_id, endpoint, run_key)
                );

                CREATE INDEX IF NOT EXISTS idx_mapping_entities_normalized_label
                    ON mapping_entities(normalized_label);
                CREATE INDEX IF NOT EXISTS idx_mapping_links_from_normalized_label
                    ON mapping_links(from_normalized_label);
                CREATE INDEX IF NOT EXISTS idx_mapping_links_to_normalized_label
                    ON mapping_links(to_normalized_label);
                CREATE INDEX IF NOT EXISTS idx_mapping_matches_run_key
                    ON mapping_matches(run_key);
                """
            )
            _ensure_text_column(
                connection,
                table_name="mapping_evidence",
                column_name="document_summary",
            )

    def clear_all(self) -> None:
        with self.managed_connection() as connection:
            connection.execute("DELETE FROM mapping_matches")
            connection.execute("DELETE FROM mapping_evidence")
            connection.execute("DELETE FROM mapping_links")
            connection.execute("DELETE FROM mapping_entities")
            connection.execute("DELETE FROM mapping_imports")

    def create_import(self, source_dir: Path) -> int:
        with self.managed_connection() as connection:
            cursor = connection.execute(
                "INSERT INTO mapping_imports(source_dir) VALUES(?)",
                (str(source_dir),),
            )
            return int(cursor.lastrowid)

    def insert_entity(
        self,
        *,
        import_id: int,
        workbook_name: str,
        sheet_name: str,
        row_number: int,
        label: str,
        entity_type: str,
        description: str,
        raw_row: list[str],
    ) -> int:
        with self.managed_connection() as connection:
            cursor = connection.execute(
                """
                INSERT INTO mapping_entities(
                    import_id,
                    workbook_name,
                    sheet_name,
                    row_number,
                    label,
                    normalized_label,
                    entity_type,
                    description,
                    raw_row_json
                ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    import_id,
                    workbook_name,
                    sheet_name,
                    row_number,
                    _clean_mapping_scalar(label),
                    normalize_mapping_label(label),
                    canonicalize_entity_type(entity_type),
                    _clean_mapping_block(description),
                    json.dumps(raw_row, ensure_ascii=False),
                ),
            )
            return int(cursor.lastrowid)

    def insert_link(
        self,
        *,
        import_id: int,
        workbook_name: str,
        sheet_name: str,
        row_number: int,
        from_label: str,
        to_label: str,
        link_type: str,
        description: str,
        raw_row: list[str],
    ) -> int:
        with self.managed_connection() as connection:
            cursor = connection.execute(
                """
                INSERT INTO mapping_links(
                    import_id,
                    workbook_name,
                    sheet_name,
                    row_number,
                    from_label,
                    from_normalized_label,
                    to_label,
                    to_normalized_label,
                    link_type,
                    description,
                    quality,
                    raw_row_json
                ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'low', ?)
                """,
                (
                    import_id,
                    workbook_name,
                    sheet_name,
                    row_number,
                    _clean_mapping_scalar(from_label),
                    normalize_mapping_label(from_label),
                    _clean_mapping_scalar(to_label),
                    normalize_mapping_label(to_label),
                    canonicalize_link_type(link_type),
                    _clean_mapping_block(description),
                    json.dumps(raw_row, ensure_ascii=False),
                ),
            )
            return int(cursor.lastrowid)

    def insert_evidence(
        self,
        *,
        mapping_link_id: int,
        ordinal: int,
        evidence_kind: str,
        title: str,
        url: str,
        snippet: str,
        document_summary: str = "",
    ) -> None:
        with self.managed_connection() as connection:
            connection.execute(
                """
                INSERT OR REPLACE INTO mapping_evidence(
                    mapping_link_id,
                    ordinal,
                    evidence_kind,
                    title,
                    url,
                    snippet,
                    document_summary
                ) VALUES(?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    mapping_link_id,
                    ordinal,
                    evidence_kind,
                    title,
                    url,
                    snippet,
                    document_summary,
                ),
            )

    def clear_matches(self, run_key: str) -> None:
        with self.managed_connection() as connection:
            connection.execute("DELETE FROM mapping_matches WHERE run_key = ?", (run_key,))

    def insert_match(
        self,
        *,
        mapping_link_id: int,
        endpoint: str,
        run_key: str,
        matched_node_id: str,
        matched_node_label: str,
        match_type: str,
    ) -> None:
        with self.managed_connection() as connection:
            connection.execute(
                """
                INSERT OR REPLACE INTO mapping_matches(
                    mapping_link_id,
                    endpoint,
                    run_key,
                    matched_node_id,
                    matched_node_label,
                    match_type
                ) VALUES(?, ?, ?, ?, ?, ?)
                """,
                (
                    mapping_link_id,
                    endpoint,
                    run_key,
                    matched_node_id,
                    matched_node_label,
                    match_type,
                ),
            )

    def list_entities(self) -> list[sqlite3.Row]:
        with self.managed_connection() as connection:
            return connection.execute(
                """
                SELECT *
                FROM mapping_entities
                ORDER BY workbook_name, sheet_name, row_number
                """
            ).fetchall()

    def list_links(self) -> list[sqlite3.Row]:
        with self.managed_connection() as connection:
            return connection.execute(
                """
                SELECT *
                FROM mapping_links
                ORDER BY workbook_name, sheet_name, row_number
                """
            ).fetchall()

    def list_evidence(self) -> list[sqlite3.Row]:
        with self.managed_connection() as connection:
            return connection.execute(
                """
                SELECT *
                FROM mapping_evidence
                ORDER BY mapping_link_id, ordinal
                """
            ).fetchall()

    def list_matches(self, run_key: str) -> list[sqlite3.Row]:
        with self.managed_connection() as connection:
            return connection.execute(
                """
                SELECT *
                FROM mapping_matches
                WHERE run_key = ?
                ORDER BY mapping_link_id, endpoint
                """,
                (run_key,),
            ).fetchall()


def import_mapping_workbooks(mapping_dir: Path, database_path: Path) -> dict[str, int]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - guarded by runtime dependency
        raise RuntimeError("openpyxl is required to import Mapping spreadsheets.") from exc

    mapping_path = Path(mapping_dir)
    workbook_paths = sorted(path for path in mapping_path.glob("*.xlsx") if path.is_file())
    store = MappingStore(database_path)
    store.init_db()
    store.clear_all()
    import_id = store.create_import(mapping_path)

    entity_count = 0
    link_count = 0
    evidence_count = 0

    with store.managed_connection() as connection:
        for workbook_path in workbook_paths:
            workbook = load_workbook(workbook_path, read_only=False, data_only=True)
            for worksheet in workbook.worksheets:
                row_iter = worksheet.iter_rows(values_only=True)
                try:
                    header_row = next(row_iter)
                except StopIteration:
                    continue
                headers = [_cell_text(value) for value in header_row]
                entity_cols = _entity_columns(headers)
                link_cols = _link_columns(headers)

                for row_index, raw_row in enumerate(row_iter, start=2):
                    values = [_cell_text(value) for value in raw_row]
                    if not any(values):
                        continue

                    if entity_cols["label"] is not None:
                        label = _clean_mapping_scalar(values[entity_cols["label"]])
                        entity_type = canonicalize_entity_type(_column_value(values, entity_cols["type"]))
                        description = _clean_mapping_block(_column_value(values, entity_cols["description"]))
                        if label:
                            cursor = connection.execute(
                                """
                                INSERT OR IGNORE INTO mapping_entities(
                                    import_id,
                                    workbook_name,
                                    sheet_name,
                                    row_number,
                                    label,
                                    normalized_label,
                                    entity_type,
                                    description,
                                    raw_row_json
                                ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """,
                                (
                                    import_id,
                                    workbook_path.name,
                                    worksheet.title,
                                    row_index,
                                    label,
                                    normalize_mapping_label(label),
                                    entity_type,
                                    description,
                                    json.dumps(values, ensure_ascii=False),
                                ),
                            )
                            if int(cursor.rowcount or 0) > 0:
                                entity_count += 1

                    from_label = _clean_mapping_scalar(_column_value(values, link_cols["from"]))
                    to_label = _clean_mapping_scalar(_column_value(values, link_cols["to"]))
                    if from_label and to_label:
                        link_type = canonicalize_link_type(_column_value(values, link_cols["type"]))
                        description = _clean_mapping_block(_column_value(values, link_cols["description"]))
                        cursor = connection.execute(
                            """
                            INSERT OR IGNORE INTO mapping_links(
                                import_id,
                                workbook_name,
                                sheet_name,
                                row_number,
                                from_label,
                                from_normalized_label,
                                to_label,
                                to_normalized_label,
                                link_type,
                                description,
                                quality,
                                raw_row_json
                            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'low', ?)
                            """,
                            (
                                import_id,
                                workbook_path.name,
                                worksheet.title,
                                row_index,
                                from_label,
                                normalize_mapping_label(from_label),
                                to_label,
                                normalize_mapping_label(to_label),
                                link_type,
                                description,
                                json.dumps(values, ensure_ascii=False),
                            ),
                        )
                        mapping_link_row = connection.execute(
                            """
                            SELECT id
                            FROM mapping_links
                            WHERE workbook_name = ? AND sheet_name = ? AND row_number = ?
                            """,
                            (
                                workbook_path.name,
                                worksheet.title,
                                row_index,
                            ),
                        ).fetchone()
                        if mapping_link_row is None:
                            continue
                        mapping_link_id = int(mapping_link_row["id"])
                        if int(cursor.rowcount or 0) > 0:
                            link_count += 1
                        for ordinal, evidence in enumerate(
                            extract_evidence_links(description), start=1
                        ):
                            evidence_cursor = connection.execute(
                                """
                                INSERT OR REPLACE INTO mapping_evidence(
                                    mapping_link_id,
                                    ordinal,
                                    evidence_kind,
                                    title,
                                    url,
                                    snippet
                                ) VALUES(?, ?, ?, ?, ?, ?)
                                """,
                                (
                                    mapping_link_id,
                                    ordinal,
                                    evidence["kind"],
                                    evidence["title"],
                                    evidence["url"],
                                    description,
                                ),
                            )
                            if int(evidence_cursor.rowcount or 0) > 0:
                                evidence_count += 1

    return {
        "workbook_count": len(workbook_paths),
        "entity_count": entity_count,
        "link_count": link_count,
        "evidence_count": evidence_count,
    }


def build_low_confidence_overlay(
    *,
    main_data: dict[str, Any],
    database_path: Path,
    run_key: str,
    include_unmatched: bool = False,
    include_generated_links: bool = False,
    enable_ai_org_matching: bool = False,
    settings: Settings | None = None,
    organisation_match_resolver: Any | None = None,
) -> dict[str, Any]:
    store = MappingStore(database_path)
    store.init_db()

    node_index: dict[str, list[dict[str, str]]] = {}
    person_index: dict[str, list[dict[str, str]]] = {}
    organisation_entries: list[dict[str, Any]] = []
    seed_identity_ids: dict[str, str] = {}
    identity_to_seed: dict[str, str] = {}
    seed_label_by_id: dict[str, str] = {}
    for edge in main_data.get("edges") or []:
        if str(edge.get("kind") or "") != "alias":
            continue
        source_id = str(edge.get("source") or "")
        target_id = str(edge.get("target") or "")
        if source_id.startswith("seed:") and target_id:
            seed_identity_ids.setdefault(source_id, target_id)
            identity_to_seed.setdefault(target_id, source_id)
        elif target_id.startswith("seed:") and source_id:
            seed_identity_ids.setdefault(target_id, source_id)
            identity_to_seed.setdefault(source_id, target_id)

    for node in main_data.get("nodes") or []:
        node_id = str(node.get("id") or "")
        label = str(node.get("label") or "").strip()
        aliases = [str(alias or "").strip() for alias in (node.get("aliases") or []) if str(alias or "").strip()]
        if str(node.get("kind") or "") == "seed" and label:
            seed_label_by_id[node_id] = label
        if str(node.get("kind") or "") != "seed":
            if label:
                _add_match_candidate(
                    node_index,
                    normalize_mapping_label(label),
                    node_id=node_id,
                    node_label=label,
                    match_type="exact_label",
                )
                if str(node.get("kind") or "") == "organisation":
                    for variant_key in _organisation_match_keys(label):
                        if variant_key == normalize_mapping_label(label):
                            continue
                        _add_match_candidate(
                            node_index,
                            variant_key,
                            node_id=node_id,
                            node_label=label,
                            match_type="organisation_variant",
                        )
            for alias_text in aliases:
                _add_match_candidate(
                    node_index,
                    normalize_mapping_label(alias_text),
                    node_id=node_id,
                    node_label=label or alias_text,
                    match_type="exact_alias",
                )
                if str(node.get("kind") or "") == "organisation":
                    for variant_key in _organisation_match_keys(alias_text):
                        if variant_key == normalize_mapping_label(alias_text):
                            continue
                        _add_match_candidate(
                            node_index,
                            variant_key,
                            node_id=node_id,
                            node_label=label or alias_text,
                            match_type="organisation_variant",
                        )
        if str(node.get("kind") or "") == "organisation" and label:
            organisation_entries.append(
                {
                    "node_id": node_id,
                    "node_label": label,
                    "registry_type": str(node.get("registry_type") or ""),
                    "match_keys": sorted({
                        key
                        for text in [label, *aliases]
                        for key in _organisation_match_keys(text)
                    }),
                }
            )
        person_like = bool(node.get("lane") in {1, 4} or str(node.get("kind") or "") in {"person", "seed_alias"})
        if person_like:
            for text in [label, *aliases]:
                for variant in _person_variant_texts(text):
                    _add_match_candidate(
                        person_index,
                        _person_match_key(variant),
                        node_id=node_id,
                        node_label=label or variant,
                        match_type="person_variant",
                    )
        if str(node.get("kind") or "") == "seed":
            identity_id = seed_identity_ids.get(node_id, "")
            if not identity_id:
                continue
            for text in [label, *aliases]:
                for variant in _person_variant_texts(text):
                    _add_match_candidate(
                        person_index,
                        _person_match_key(variant),
                        node_id=identity_id,
                        node_label=label or variant,
                        match_type="seed_alias_variant",
                    )

    entity_rows = store.list_entities()
    entity_by_normalized: dict[str, sqlite3.Row] = {}
    for row in entity_rows:
        entity_by_normalized.setdefault(str(row["normalized_label"]), row)

    evidence_rows = store.list_evidence()
    evidence_by_link_id: dict[int, list[sqlite3.Row]] = {}
    for row in evidence_rows:
        evidence_by_link_id.setdefault(int(row["mapping_link_id"]), []).append(row)

    store.clear_matches(run_key)

    if organisation_match_resolver is None and enable_ai_org_matching:
        effective_settings = settings or load_settings()
        organisation_match_resolver = OrganisationMatchResolver(effective_settings)

    overlay_nodes: dict[str, dict[str, Any]] = {}
    overlay_edges: list[dict[str, Any]] = []
    matched_link_count = 0
    generated_doc_nodes_by_sheet: dict[str, set[str]] = {}
    generated_doc_nodes_by_signer: dict[str, dict[str, set[str]]] = {}
    generated_affiliations: list[dict[str, Any]] = []

    for link_row in store.list_links():
        if (
            not include_generated_links
            and str(link_row["workbook_name"] or "").strip() == "__evidence_enrichment__"
        ):
            continue
        from_entity_row = entity_by_normalized.get(str(link_row["from_normalized_label"]))
        to_entity_row = entity_by_normalized.get(str(link_row["to_normalized_label"]))
        from_match = _unique_match(node_index, str(link_row["from_normalized_label"]))
        to_match = _unique_match(node_index, str(link_row["to_normalized_label"]))
        from_entity_type = canonicalize_entity_type(str(from_entity_row["entity_type"] or "").strip()) if from_entity_row else ""
        to_entity_type = canonicalize_entity_type(str(to_entity_row["entity_type"] or "").strip()) if to_entity_row else ""
        if not from_match and from_entity_type == "individual":
            from_match = _unique_match(person_index, _person_match_key(str(link_row["from_label"])))
        if not to_match and to_entity_type == "individual":
            to_match = _unique_match(person_index, _person_match_key(str(link_row["to_label"])))
        link_type = canonicalize_link_type(str(link_row["link_type"] or ""))
        if (
            not from_match
            and _should_try_ai_organisation_match(
                entity_type=from_entity_type,
                label=str(link_row["from_label"]),
                link_type=link_type,
                endpoint="from",
            )
        ):
            from_match = _resolve_organisation_match_with_ai(
                label=str(link_row["from_label"]),
                link_type=link_type,
                endpoint="from",
                counterpart_label=str(link_row["to_label"]),
                description=str(link_row["description"] or ""),
                organisation_entries=organisation_entries,
                organisation_match_resolver=organisation_match_resolver,
                workbook_name=str(link_row["workbook_name"] or ""),
                sheet_name=str(link_row["sheet_name"] or ""),
                row_number=int(link_row["row_number"]),
            )
        if (
            not to_match
            and _should_try_ai_organisation_match(
                entity_type=to_entity_type,
                label=str(link_row["to_label"]),
                link_type=link_type,
                endpoint="to",
            )
        ):
            to_match = _resolve_organisation_match_with_ai(
                label=str(link_row["to_label"]),
                link_type=link_type,
                endpoint="to",
                counterpart_label=str(link_row["from_label"]),
                description=str(link_row["description"] or ""),
                organisation_entries=organisation_entries,
                organisation_match_resolver=organisation_match_resolver,
                workbook_name=str(link_row["workbook_name"] or ""),
                sheet_name=str(link_row["sheet_name"] or ""),
                row_number=int(link_row["row_number"]),
            )

        if from_match:
            store.insert_match(
                mapping_link_id=int(link_row["id"]),
                endpoint="from",
                run_key=run_key,
                matched_node_id=from_match["node_id"],
                matched_node_label=from_match["node_label"],
                match_type=from_match["match_type"],
            )
        if to_match:
            store.insert_match(
                mapping_link_id=int(link_row["id"]),
                endpoint="to",
                run_key=run_key,
                matched_node_id=to_match["node_id"],
                matched_node_label=to_match["node_label"],
                match_type=to_match["match_type"],
            )

        if not include_unmatched and not (from_match or to_match):
            continue

        evidence_items = [
            {
                "title": str(item["title"] or item["url"] or "Evidence").strip(),
                "document_url": str(item["url"] or "").strip(),
                "page_hint": "",
                "page_number": None,
                "notes": summarize_mapping_text(
                    str(item["document_summary"] or item["snippet"] or "").strip()
                ),
            }
            for item in evidence_by_link_id.get(int(link_row["id"]), [])
            if str(item["url"] or "").strip()
        ]
        summary_text = next(
            (
                str(item["document_summary"] or "").strip()
                for item in evidence_by_link_id.get(int(link_row["id"]), [])
                if str(item["document_summary"] or "").strip()
            ),
            "",
        )
        description = summary_text or str(link_row["description"] or "").strip()
        tooltip_description = summarize_mapping_text(description)

        source_id = from_match["node_id"] if from_match else _ensure_overlay_node(
            overlay_nodes,
            label=str(link_row["from_label"]),
            normalized_label=str(link_row["from_normalized_label"]),
            entity_row=from_entity_row,
            link_type=link_type,
            endpoint="from",
            description_hint=description,
        )
        target_id = to_match["node_id"] if to_match else _ensure_overlay_node(
            overlay_nodes,
            label=str(link_row["to_label"]),
            normalized_label=str(link_row["to_normalized_label"]),
            entity_row=to_entity_row,
            link_type=link_type,
            endpoint="to",
            description_hint=description,
        )
        if source_id == target_id:
            continue
        phrase = _mapping_phrase(link_type)
        tooltip_lines = [
            line
            for line in [link_type, tooltip_description]
            if line
        ]
        tooltip_lines.append(
            f"Imported from {link_row['workbook_name']} / {link_row['sheet_name']} / row {link_row['row_number']}"
        )
        source_node = overlay_nodes.get(source_id) if source_id in overlay_nodes else next(
            (node for node in main_data.get("nodes") or [] if str(node.get("id")) == source_id),
            None,
        )
        target_node = overlay_nodes.get(target_id) if target_id in overlay_nodes else next(
            (node for node in main_data.get("nodes") or [] if str(node.get("id")) == target_id),
            None,
        )
        workbook_name = str(link_row["workbook_name"] or "").strip()
        sheet_name = str(link_row["sheet_name"] or "").strip()
        if workbook_name == "__evidence_enrichment__":
            if "signatory" in link_type:
                signer_id = ""
                document_id = ""
                if (
                    source_node
                    and _is_person_anchor_node(source_node)
                    and target_node
                    and bool(target_node.get("low_confidence_expandable"))
                ):
                    signer_id = source_id
                    document_id = target_id
                elif (
                    target_node
                    and _is_person_anchor_node(target_node)
                    and source_node
                    and bool(source_node.get("low_confidence_expandable"))
                ):
                    signer_id = target_id
                    document_id = source_id
                if signer_id and document_id:
                    generated_doc_nodes_by_sheet.setdefault(sheet_name, set()).add(document_id)
                    signer_docs = generated_doc_nodes_by_signer.setdefault(sheet_name, {})
                    signer_docs.setdefault(signer_id, set()).add(document_id)
            elif "affiliate" in link_type:
                signer_id = ""
                organisation_id = ""
                if source_node and _is_person_anchor_node(source_node) and target_node and str(target_node.get("kind") or "") == "organisation":
                    signer_id = source_id
                    organisation_id = target_id
                elif target_node and _is_person_anchor_node(target_node) and source_node and str(source_node.get("kind") or "") == "organisation":
                    signer_id = target_id
                    organisation_id = source_id
                if signer_id and organisation_id:
                    generated_affiliations.append(
                        {
                            "sheet_name": sheet_name,
                            "signer_id": signer_id,
                            "organisation_id": organisation_id,
                            "source_id": source_id,
                            "target_id": target_id,
                            "phrase": phrase,
                            "link_type": link_type,
                            "organisation_label": str(target_node.get("label") if organisation_id == target_id else source_node.get("label") if source_node else ""),
                            "tooltip_description": tooltip_description,
                            "evidence_items": evidence_items,
                            "tooltip_lines": tooltip_lines,
                            "link_row_id": int(link_row["id"]),
                        }
                    )
                    continue
        overlay_edges.append(
            {
                "id": f"mapping-link:{int(link_row['id'])}",
                "source": source_id,
                "target": target_id,
                "kind": "mapping_link",
                "phrase": phrase,
                "role_type": link_type or "mapping_link",
                "role_label": link_type or "mapping_link",
                "source_provider": "mapping_import",
                "confidence": "low",
                "weight": 0.2,
                "tooltip": tooltip_description
                or f"{link_row['from_label']} {phrase} {link_row['to_label']}",
                "tooltip_lines": tooltip_lines,
                "is_low_confidence": True,
                "evidence": evidence_items[0] if evidence_items else None,
                "evidence_items": evidence_items,
            }
        )
        if from_match or to_match:
            matched_link_count += 1

    seen_derived_edge_keys: set[tuple[str, str, str]] = set()
    for affiliation in generated_affiliations:
        doc_ids = generated_doc_nodes_by_signer.get(affiliation["sheet_name"], {}).get(affiliation["signer_id"], set())
        if not doc_ids:
            direct_key = (affiliation["source_id"], affiliation["target_id"], affiliation["link_type"])
            if direct_key in seen_derived_edge_keys:
                continue
            seen_derived_edge_keys.add(direct_key)
            overlay_edges.append(
                {
                    "id": f"mapping-link:{affiliation['link_row_id']}",
                    "source": affiliation["source_id"],
                    "target": affiliation["target_id"],
                    "kind": "mapping_link",
                    "phrase": affiliation["phrase"],
                    "role_type": affiliation["link_type"] or "mapping_link",
                    "role_label": affiliation["link_type"] or "mapping_link",
                    "source_provider": "mapping_import",
                    "confidence": "low",
                    "weight": 0.2,
                    "tooltip": affiliation["tooltip_description"]
                    or f"{affiliation['source_id']} {affiliation['phrase']} {affiliation['organisation_label']}",
                    "tooltip_lines": affiliation["tooltip_lines"],
                    "is_low_confidence": True,
                    "evidence": affiliation["evidence_items"][0] if affiliation["evidence_items"] else None,
                    "evidence_items": affiliation["evidence_items"],
                }
            )
            continue
        for document_id in doc_ids:
            derived_key = (document_id, affiliation["organisation_id"], "represented_organisation")
            if derived_key in seen_derived_edge_keys:
                continue
            seen_derived_edge_keys.add(derived_key)
            overlay_edges.append(
                {
                    "id": f"mapping-doc-link:{slugify_mapping_label(document_id)}:{slugify_mapping_label(affiliation['organisation_id'])}",
                    "source": document_id,
                    "target": affiliation["organisation_id"],
                    "kind": "mapping_document_affiliation",
                    "phrase": "includes signatories representing",
                    "role_type": "represented_organisation",
                    "role_label": "represented organisation",
                    "source_provider": "mapping_import",
                    "confidence": "low",
                    "weight": 0.2,
                    "tooltip": affiliation["tooltip_description"] or f"{document_id} includes signatories representing {affiliation['organisation_label']}",
                    "tooltip_lines": [
                        "includes signatories representing",
                        *[
                            line
                            for line in affiliation["tooltip_lines"][1:]
                            if line and not line.startswith("Imported from ")
                        ],
                    ],
                    "is_low_confidence": True,
                    "evidence": affiliation["evidence_items"][0] if affiliation["evidence_items"] else None,
                    "evidence_items": affiliation["evidence_items"],
                }
            )

    return {
        "nodes": list(overlay_nodes.values()),
        "edges": overlay_edges,
        "summary": {
            "run_key": run_key,
            "overlay_node_count": len(overlay_nodes),
            "overlay_edge_count": len(overlay_edges),
            "matched_link_count": matched_link_count,
        },
    }


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _column_value(values: list[str], index: int | None) -> str:
    if index is None or index >= len(values):
        return ""
    return values[index]


def _find_header_index(headers: list[str], target: str, *, start: int = 0) -> int | None:
    for index in range(start, len(headers)):
        if headers[index].strip().lower() == target:
            return index
    return None


def _entity_columns(headers: list[str]) -> dict[str, int | None]:
    label_index = _find_header_index(headers, "label")
    if label_index is None:
        return {"label": None, "type": None, "description": None}
    return {
        "label": label_index,
        "type": _find_header_index(headers, "type", start=label_index + 1),
        "description": _find_header_index(headers, "description", start=label_index + 1),
    }


def _link_columns(headers: list[str]) -> dict[str, int | None]:
    from_index = _find_header_index(headers, "from")
    to_index = _find_header_index(headers, "to")
    if from_index is None or to_index is None:
        return {"from": None, "to": None, "type": None, "description": None}
    start = max(from_index, to_index) + 1
    return {
        "from": from_index,
        "to": to_index,
        "type": _find_header_index(headers, "type", start=start),
        "description": _find_header_index(headers, "description", start=start),
    }


def _unique_match(
    node_index: dict[str, list[dict[str, str]]],
    normalized_label: str,
) -> dict[str, str] | None:
    matches = node_index.get(normalized_label, [])
    unique_by_node_id: dict[str, dict[str, str]] = {}
    for match in matches:
        unique_by_node_id.setdefault(match["node_id"], match)
    if len(unique_by_node_id) != 1:
        return None
    return next(iter(unique_by_node_id.values()))


def _person_match_key(value: str) -> str:
    return normalize_mapping_label(normalize_name(value))


def _add_match_candidate(
    index: dict[str, list[dict[str, str]]],
    key: str,
    *,
    node_id: str,
    node_label: str,
    match_type: str,
) -> None:
    normalized_key = str(key or "").strip()
    if not normalized_key:
        return
    index.setdefault(normalized_key, []).append(
        {
            "node_id": node_id,
            "node_label": node_label,
            "match_type": match_type,
        }
    )


def _person_variant_texts(value: str) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    variants = [text]
    try:
        variants.extend(item.name for item in generate_name_variants(text, "balanced"))
    except Exception:
        pass
    return list(dict.fromkeys(item for item in variants if str(item or "").strip()))


def _organisation_match_keys(value: str) -> list[str]:
    normalized = normalize_mapping_label(value)
    if not normalized:
        return []
    keys = [normalized]
    tokens = normalized.split()
    while tokens and tokens[-1] in _ORGANISATION_SUFFIX_TOKENS:
        tokens = tokens[:-1]
        if tokens:
            keys.append(" ".join(tokens))
    return list(dict.fromkeys(key for key in keys if key))


def _is_organisation_entity_type(entity_type: str) -> bool:
    lowered = str(entity_type or "").strip().lower()
    return (
        "organisation" in lowered
        or "organization" in lowered
        or "company" in lowered
        or "charity" in lowered
        or lowered == "org"
    )


def _looks_like_overlay_document(label: str, *, link_type: str, endpoint: str) -> bool:
    normalized = normalize_mapping_label(label)
    if any(hint in normalized for hint in _OTHER_ORGANISATION_HINTS):
        return True
    return endpoint == "to" and "signatory" in str(link_type or "").strip().lower()


def _should_try_ai_organisation_match(
    *,
    entity_type: str,
    label: str,
    link_type: str,
    endpoint: str,
) -> bool:
    return (
        _is_organisation_entity_type(entity_type)
        and bool(str(label or "").strip())
        and not _looks_like_overlay_document(label, link_type=link_type, endpoint=endpoint)
    )


def _score_organisation_candidate(
    label: str,
    candidate: dict[str, Any],
) -> dict[str, Any] | None:
    normalized_label = normalize_mapping_label(label)
    label_tokens = set(normalized_label.split())
    if not normalized_label or len(label_tokens) < 2:
        return None
    best: dict[str, Any] | None = None
    for key in candidate.get("match_keys") or []:
        key_tokens = set(str(key or "").split())
        overlap = len(label_tokens & key_tokens)
        subset = bool(label_tokens and (label_tokens.issubset(key_tokens) or key_tokens.issubset(label_tokens)))
        ratio = SequenceMatcher(None, normalized_label, str(key or "")).ratio()
        if overlap < _ORGANISATION_AI_MIN_OVERLAP:
            continue
        if not subset and ratio < _ORGANISATION_AI_MIN_RATIO:
            continue
        scored = {
            "node_id": candidate["node_id"],
            "node_label": candidate["node_label"],
            "registry_type": candidate.get("registry_type", ""),
            "matched_key": str(key or ""),
            "overlap": overlap,
            "subset": subset,
            "ratio": ratio,
            "score": (100.0 if subset else 0.0) + (overlap * 10.0) + ratio,
        }
        if best is None or scored["score"] > best["score"]:
            best = scored
    return best


def _organisation_ai_candidates(
    label: str,
    organisation_entries: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    scored = []
    for entry in organisation_entries:
        candidate = _score_organisation_candidate(label, entry)
        if candidate:
            scored.append(candidate)
    scored.sort(
        key=lambda item: (-float(item["score"]), -int(item["overlap"]), -float(item["ratio"]), str(item["node_label"])),
    )
    seen_node_ids: set[str] = set()
    unique_candidates: list[dict[str, Any]] = []
    for item in scored:
        if item["node_id"] in seen_node_ids:
            continue
        seen_node_ids.add(item["node_id"])
        unique_candidates.append(item)
        if len(unique_candidates) >= _ORGANISATION_AI_MAX_CANDIDATES:
            break
    return unique_candidates


@dataclass(slots=True)
class OrganisationMatchResolver:
    settings: Settings
    _gemini: GeminiClient | None = field(init=False, default=None)
    _openai: OpenAIResponsesClient | None = field(init=False, default=None)

    def __post_init__(self) -> None:
        if self.settings.resolution_provider == "gemini" and self.settings.gemini_api_key:
            self._gemini = GeminiClient(
                api_key=self.settings.gemini_api_key,
                cache_dir=self.settings.cache_dir / "gemini_low_confidence_org_resolution",
            )
        elif self.settings.openai_api_key:
            self._openai = OpenAIResponsesClient(
                api_key=self.settings.openai_api_key,
                base_url=self.settings.openai_base_url,
                cache_dir=self.settings.cache_dir / "openai_low_confidence_org_resolution",
                user_agent=self.settings.user_agent,
            )

    @property
    def has_llm(self) -> bool:
        return self._gemini is not None or self._openai is not None

    def resolve(
        self,
        *,
        label: str,
        candidates: list[dict[str, Any]],
        link_type: str,
        endpoint: str,
        counterpart_label: str,
        description: str,
        workbook_name: str,
        sheet_name: str,
        row_number: int,
    ) -> dict[str, str] | None:
        if not self.has_llm or not candidates:
            return None
        prompt = _build_organisation_resolution_prompt(
            label=label,
            candidates=candidates,
            link_type=link_type,
            endpoint=endpoint,
            counterpart_label=counterpart_label,
            description=description,
            workbook_name=workbook_name,
            sheet_name=sheet_name,
            row_number=row_number,
        )
        try:
            if self._gemini is not None:
                response = self._gemini.generate(
                    model=self.settings.gemini_resolution_model,
                    prompt=prompt,
                )
                document = extract_json_document(extract_gemini_text(response))
            else:
                response = self._openai.create_response(
                    model=self.settings.openai_resolution_model,
                    input_text=prompt,
                    metadata={"task": "low_confidence_org_resolution"},
                )
                document = extract_json_document(extract_output_text(response))
        except Exception as exc:
            log.warning("Low-confidence organisation LLM match failed for %r: %s", label, exc)
            return None
        if str(document.get("status") or "").strip().lower() != "match":
            return None
        candidate_id = str(document.get("candidate_id") or "").strip()
        match = next((candidate for candidate in candidates if candidate["node_id"] == candidate_id), None)
        if not match:
            return None
        return {
            "node_id": match["node_id"],
            "node_label": match["node_label"],
            "match_type": "organisation_ai",
        }


def _build_organisation_resolution_prompt(
    *,
    label: str,
    candidates: list[dict[str, Any]],
    link_type: str,
    endpoint: str,
    counterpart_label: str,
    description: str,
    workbook_name: str,
    sheet_name: str,
    row_number: int,
) -> str:
    candidate_lines = []
    for candidate in candidates:
        candidate_lines.append(
            "\n".join(
                [
                    f"- candidate_id: {candidate['node_id']}",
                    f"  label: {candidate['node_label']}",
                    f"  registry_type: {candidate.get('registry_type', '')}",
                    f"  matched_key: {candidate.get('matched_key', '')}",
                    f"  token_overlap: {candidate.get('overlap', 0)}",
                    f"  subset_match: {str(candidate.get('subset', False)).lower()}",
                    f"  similarity: {round(float(candidate.get('ratio') or 0.0), 3)}",
                ]
            )
        )
    return f"""\
Choose whether the imported low-confidence organisation label matches one existing organisation node in the main graph.
Return JSON only with this shape:
{{
  "status": "match" | "no_match",
  "candidate_id": "",
  "explanation": ""
}}

Imported organisation label: {label}
Endpoint: {endpoint}
Counterpart label: {counterpart_label}
Link type: {link_type}
Workbook: {workbook_name}
Sheet: {sheet_name}
Row: {row_number}
Description: {summarize_mapping_text(description, max_chars=300)}

Match only if the candidate is the same real-world organisation, not merely a thematically similar body.

Candidates:
{chr(10).join(candidate_lines)}
"""


def _resolve_organisation_match_with_ai(
    *,
    label: str,
    link_type: str,
    endpoint: str,
    counterpart_label: str,
    description: str,
    organisation_entries: list[dict[str, Any]],
    organisation_match_resolver: Any | None,
    workbook_name: str,
    sheet_name: str,
    row_number: int,
) -> dict[str, str] | None:
    if organisation_match_resolver is None:
        return None
    candidates = _organisation_ai_candidates(label, organisation_entries)
    if not candidates:
        return None
    return organisation_match_resolver.resolve(
        label=label,
        candidates=candidates,
        link_type=link_type,
        endpoint=endpoint,
        counterpart_label=counterpart_label,
        description=description,
        workbook_name=workbook_name,
        sheet_name=sheet_name,
        row_number=row_number,
    )


def _promote_person_match_to_seed(
    match: dict[str, str] | None,
    *,
    identity_to_seed: dict[str, str],
    seed_label_by_id: dict[str, str],
) -> dict[str, str] | None:
    if not match:
        return None
    seed_id = identity_to_seed.get(str(match.get("node_id") or ""))
    if not seed_id:
        return match
    return {
        "node_id": seed_id,
        "node_label": seed_label_by_id.get(seed_id) or str(match.get("node_label") or seed_id),
        "match_type": f"{match.get('match_type', 'match')}_seed",
    }


def _infer_overlay_node_kind(
    entity_type: str,
    *,
    link_type: str = "",
    endpoint: str = "",
) -> tuple[str, int, str, bool]:
    lowered = str(entity_type or "").strip().lower()
    lowered_link_type = str(link_type or "").strip().lower()
    if "address" in lowered:
        return ("address", 3, "", False)
    if lowered in {"individual", "person"} or "individual" in lowered or "person" in lowered:
        return ("person", 4, "", False)
    if "charity" in lowered:
        return ("organisation", 2, "charity", False)
    if "company" in lowered:
        return ("organisation", 2, "company", False)
    if any(token in lowered for token in _OTHER_ORGANISATION_HINTS):
        return ("organisation", 2, "other", True)
    if endpoint == "to" and "signatory" in lowered_link_type:
        return ("organisation", 2, "other", True)
    return ("organisation", 2, "", False)


def _ensure_overlay_node(
    overlay_nodes: dict[str, dict[str, Any]],
    *,
    label: str,
    normalized_label: str,
    entity_row: sqlite3.Row | None,
    link_type: str = "",
    endpoint: str = "",
    description_hint: str = "",
) -> str:
    node_id = f"mapping-node:{slugify_mapping_label(normalized_label or label)}"
    if node_id in overlay_nodes:
        return node_id

    entity_type = canonicalize_entity_type(str(entity_row["entity_type"] or "").strip()) if entity_row else ""
    description = (
        str(entity_row["description"] or "").strip()
        if entity_row
        else str(description_hint or "").strip()
    )
    kind, lane, registry_type, is_expandable = _infer_overlay_node_kind(
        entity_type,
        link_type=link_type,
        endpoint=endpoint,
    )
    display_entity_type = entity_type or ("other organisation" if registry_type == "other" else "")
    tooltip_lines = [
        line for line in [display_entity_type, summarize_mapping_text(description, max_chars=160)] if line
    ]
    if entity_row is not None:
        tooltip_lines.append(
            f"Imported from {entity_row['workbook_name']} / {entity_row['sheet_name']} / row {entity_row['row_number']}"
        )
    overlay_nodes[node_id] = {
        "id": node_id,
        "label": label,
        "kind": kind,
        "lane": lane,
        "registry_type": registry_type,
        "aliases": [],
        "tooltip_lines": tooltip_lines,
        "is_low_confidence": True,
        "mapping_entity_type": display_entity_type,
        "low_confidence_expandable": is_expandable,
    }
    return node_id


def _mapping_phrase(link_type: str) -> str:
    lowered = str(link_type or "").strip().lower()
    if "signatory" in lowered:
        return "signed"
    if "affiliate" in lowered:
        return "is affiliated with"
    if "sponsor" in lowered:
        return "sponsored"
    if "spokesperson" in lowered:
        return "acted as spokesperson for"
    if "campaign" in lowered:
        return "is linked through"
    return "is linked to"


def _display_endpoint_metadata(
    *,
    label: str,
    normalized_label: str,
    entity_row: sqlite3.Row | None,
    link_type: str,
    endpoint: str,
    description_hint: str = "",
) -> tuple[str, int, str, bool]:
    entity_type = canonicalize_entity_type(str(entity_row["entity_type"] or "").strip()) if entity_row else ""
    return _infer_overlay_node_kind(
        entity_type,
        link_type=link_type,
        endpoint=endpoint,
    )


def _is_person_anchor_node(node: dict[str, Any] | None) -> bool:
    if not node:
        return False
    kind = str(node.get("kind") or "")
    lane = int(node.get("lane") or 0)
    return kind in {"person", "seed", "seed_alias"} or lane in {0, 1, 4}
